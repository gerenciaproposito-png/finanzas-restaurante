from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.models import VentaProducto, Configuracion


SHEET_NAME = "estadisticas_ventas_productos"


def parse_excel(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.worksheets[0]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        negocio, categoria, producto, cantidad, total = (row + (None,) * 5)[:5]
        if not producto:
            continue
        try:
            cant = int(round(float(cantidad))) if cantidad is not None else 0
            tot = Decimal(str(total)) if total is not None else Decimal("0")
        except (InvalidOperation, ValueError, TypeError):
            cant, tot = 0, Decimal("0")
        rows.append({
            "negocio": str(negocio).strip() if negocio else None,
            "categoria": str(categoria).strip() if categoria else "SIN CATEGORÍA",
            "producto": str(producto).strip(),
            "cantidad": cant,
            "total": tot,
        })
    return rows


def importar_lote(db: Session, rows: list[dict], fecha_corte: date, fuente: str, reemplazar: bool) -> int:
    if reemplazar:
        db.query(VentaProducto).filter(VentaProducto.fecha_corte == fecha_corte).delete()
    now = datetime.now()
    for r in rows:
        db.add(VentaProducto(
            fecha_corte=fecha_corte,
            negocio=r["negocio"],
            categoria=r["categoria"],
            producto=r["producto"],
            cantidad=r["cantidad"],
            total=r["total"],
            fuente=fuente,
            importado=now,
        ))
    db.commit()
    return len(rows)


def latest_fecha_corte(db: Session) -> date | None:
    return db.query(func.max(VentaProducto.fecha_corte)).scalar()


def get_fechas_con_conteo(db: Session) -> list[tuple[date, int]]:
    rows = (
        db.query(VentaProducto.fecha_corte, func.count(VentaProducto.id))
        .group_by(VentaProducto.fecha_corte)
        .order_by(VentaProducto.fecha_corte.desc())
        .all()
    )
    return rows


def drive_import(db: Session, fecha_corte: date, reemplazar: bool) -> dict:
    from app.services import drive

    cfg = db.get(Configuracion, "ventas_productos_folder_id")
    folder_id = cfg.valor if cfg else ""
    if not folder_id:
        return {"error": "No hay carpeta de Drive configurada para ventas por producto"}

    if not drive.is_connected():
        return {"error": "Google Drive no está conectado"}

    files = drive.list_excel_in_folder(folder_id)
    if not files:
        return {"error": "No se encontraron archivos .xlsx en la carpeta"}

    latest = files[0]
    dest_dir = Path(__file__).resolve().parent.parent.parent / "uploads" / "ventas_productos"
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / f"{latest['id']}.xlsx"
    drive.download_file(latest["id"], dest)

    rows = parse_excel(dest)
    n = importar_lote(db, rows, fecha_corte, "drive", reemplazar)
    return {"filas": n, "archivo": latest["name"]}
