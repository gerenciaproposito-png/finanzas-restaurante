import calendar
import json
from datetime import date
from decimal import Decimal

from fastapi import APIRouter, Depends, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from sqlalchemy import extract, func
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import Categoria, Gasto, Nomina, Venta

router = APIRouter(prefix="/reportes", tags=["reportes"])

UVT_2026 = Decimal("49799")  # UVT Colombia 2026 (estimado)


def _ventas_por_mes(db: Session, year: int) -> list[dict]:
    rows = (
        db.query(
            extract("month", Venta.fecha).label("mes"),
            func.coalesce(func.sum(Venta.total), 0).label("total"),
        )
        .filter(extract("year", Venta.fecha) == year)
        .group_by("mes")
        .all()
    )
    by_month = {int(r.mes): float(r.total) for r in rows}
    return [{"mes": m, "nombre": calendar.month_abbr[m], "total": by_month.get(m, 0)} for m in range(1, 13)]


def _gastos_por_mes(db: Session, year: int) -> list[dict]:
    rows = (
        db.query(
            extract("month", Gasto.fecha).label("mes"),
            func.coalesce(func.sum(Gasto.total), 0).label("total"),
        )
        .filter(extract("year", Gasto.fecha) == year)
        .group_by("mes")
        .all()
    )
    by_month = {int(r.mes): float(r.total) for r in rows}
    return [{"mes": m, "nombre": calendar.month_abbr[m], "total": by_month.get(m, 0)} for m in range(1, 13)]


def _gastos_por_categoria(db: Session, year: int, mes: int | None = None) -> list[dict]:
    q = (
        db.query(Categoria.nombre, func.coalesce(func.sum(Gasto.total), 0))
        .join(Gasto, Gasto.categoria_id == Categoria.id)
        .filter(extract("year", Gasto.fecha) == year)
    )
    if mes:
        q = q.filter(extract("month", Gasto.fecha) == mes)
    rows = q.group_by(Categoria.nombre).order_by(func.sum(Gasto.total).desc()).all()
    return [{"nombre": r[0], "total": float(r[1])} for r in rows]


@router.get("", response_class=HTMLResponse)
def vista(request: Request, db: Session = Depends(get_db), year: int = None, mes: int = None):
    hoy = date.today()
    year = year or hoy.year
    mes_actual = mes  # None = todo el año

    ventas_mes = _ventas_por_mes(db, year)
    gastos_mes = _gastos_por_mes(db, year)

    utilidad_mes = [
        {"mes": v["mes"], "nombre": v["nombre"], "utilidad": v["total"] - g["total"]}
        for v, g in zip(ventas_mes, gastos_mes)
    ]

    total_ventas_año = sum(v["total"] for v in ventas_mes)
    total_gastos_año = sum(g["total"] for g in gastos_mes)
    utilidad_año = total_ventas_año - total_gastos_año

    # Nómina del año
    total_nomina_año = db.query(func.coalesce(func.sum(Nomina.neto_pagado), 0)).filter(
        extract("year", Nomina.fecha_fin) == year
    ).scalar() or Decimal("0")

    # SIMPLE estimado (grupo 4: expendio de comidas/bebidas)
    ingresos_uvt = Decimal(str(total_ventas_año)) / UVT_2026
    from app.utils import tarifa_simple
    tarifa = tarifa_simple(ingresos_uvt)
    simple_estimado = Decimal(str(total_ventas_año)) * tarifa

    gastos_cat = _gastos_por_categoria(db, year, mes_actual)

    años_disponibles = [hoy.year, hoy.year - 1]

    return request.app.state.templates.TemplateResponse("reportes/index.html", {
        "request": request,
        "year": year,
        "mes_actual": mes_actual,
        "años_disponibles": años_disponibles,
        "ventas_mes": json.dumps([v["total"] for v in ventas_mes]),
        "gastos_mes_data": json.dumps([g["total"] for g in gastos_mes]),
        "utilidad_mes": json.dumps([u["utilidad"] for u in utilidad_mes]),
        "labels_mes": json.dumps([v["nombre"] for v in ventas_mes]),
        "total_ventas_año": total_ventas_año,
        "total_gastos_año": total_gastos_año,
        "utilidad_año": utilidad_año,
        "total_nomina_año": float(total_nomina_año),
        "tarifa_simple": float(tarifa * 100),
        "simple_estimado": float(simple_estimado),
        "gastos_cat": json.dumps([g["nombre"] for g in gastos_cat]),
        "gastos_cat_totales": json.dumps([g["total"] for g in gastos_cat]),
    })


@router.get("/exportar/ventas")
def exportar_ventas(db: Session = Depends(get_db), year: int = None):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    hoy = date.today()
    year = year or hoy.year
    ventas = db.query(Venta).filter(extract("year", Venta.fecha) == year).order_by(Venta.fecha).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Ventas {year}"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    headers = ["Fecha", "Turno", "Método pago", "Propinas", "Total"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row, v in enumerate(ventas, 2):
        ws.cell(row=row, column=1, value=str(v.fecha))
        ws.cell(row=row, column=2, value=v.turno or "")
        ws.cell(row=row, column=3, value=v.metodo_pago or "")
        ws.cell(row=row, column=4, value=float(v.propinas))
        ws.cell(row=row, column=5, value=float(v.total))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=ventas_{year}.xlsx"},
    )


@router.get("/exportar/gastos")
def exportar_gastos(db: Session = Depends(get_db), year: int = None):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    hoy = date.today()
    year = year or hoy.year
    gastos = db.query(Gasto).filter(extract("year", Gasto.fecha) == year).order_by(Gasto.fecha).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Gastos {year}"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    headers = ["Fecha", "Proveedor", "Categoría", "Descripción", "Factura", "Subtotal", "IVA", "Total"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row, g in enumerate(gastos, 2):
        ws.cell(row=row, column=1, value=str(g.fecha))
        ws.cell(row=row, column=2, value=g.proveedor.nombre if g.proveedor else "")
        ws.cell(row=row, column=3, value=g.categoria.nombre if g.categoria else "")
        ws.cell(row=row, column=4, value=g.descripcion or "")
        ws.cell(row=row, column=5, value=g.num_factura or "")
        ws.cell(row=row, column=6, value=float(g.subtotal))
        ws.cell(row=row, column=7, value=float(g.iva))
        ws.cell(row=row, column=8, value=float(g.total))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=gastos_{year}.xlsx"},
    )
